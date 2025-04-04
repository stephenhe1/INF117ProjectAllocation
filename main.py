import shutil
import re
import os
from openpyxl import load_workbook
import pandas as pd
import numpy as np
pd.set_option('display.max_columns', None)

# Corrected project mapping
project_columns_map_corrected = {
    'Rank the following projects from 1 (most preferred) to 16 (least preferred).': 'Apples and Oranges Art',
    'Unnamed: 22': 'Cashew',
    'Unnamed: 23': 'DatasafeHR',
    'Unnamed: 24': 'Daygent',
    'Unnamed: 25': 'ForOurLastNames',
    'Unnamed: 26': 'Jascot Development',
    'Unnamed: 27': 'Laguna Ocean Foundation',
    'Unnamed: 28': 'Makapo',
    'Unnamed: 29': 'MilMentor',
    'Unnamed: 30': 'ROID INC',
    'Unnamed: 31': 'Skystart App/MVP',
    'Unnamed: 32': 'SVMMARY',
    'Unnamed: 33': 'SuperNova Academy',
    'Unnamed: 34': 'Balnce.ai',
    'Unnamed: 35': 'UCI School of Education - Zhenyao Cai',
    'Unnamed: 36': 'UCI Social Ecology - Richard Matthew'
}


def calculate_project_preference(file_path):
    # Read the Excel file
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")

    # Filter the rows with the value "Completed"
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Extract the project columns
    project_columns = [col for col in completed_rows.columns if
                       any(key in col for key in project_columns_map_corrected.keys())]

    # Extract the rank values from the completed rows
    project_ranks = completed_rows[project_columns]

    # Rename the columns to use project names instead of original column names
    project_ranks.columns = [project_columns_map_corrected[col] for col in project_columns_map_corrected.keys() if
                             col in project_columns]

    # Normalize the ranks (1 = most preferred, 16 = least preferred)
    min_rank = 1
    max_rank = 16

    # Reset index for project_ranks to ensure alignment
    project_ranks.reset_index(drop=True, inplace=True)

    # Extract 'Full Name' for constructing dictionary
    student_names = completed_rows['Full Name'].reset_index(drop=True)

    # Create the dictionary
    student_preferences = {}

    # Iterate over the rows to populate the dictionary with normalized values
    for index, name in student_names.items():
        preferences = project_ranks.iloc[index].to_dict()
        # Corrected normalization formula: 1 - ((rank - min_rank) / (max_rank - min_rank))
        normalized_preferences = {project: 1 - ((rank - min_rank) / (max_rank - min_rank)) for project, rank in
                                  preferences.items()}
        student_preferences[name] = normalized_preferences

    return student_preferences


def first_come_first_serve_score(file_path):
    import pandas as pd

    # Read the Excel file
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")

    # Filter the rows with the value "Completed"
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Extract the 'Timestamp (mm/dd/yyyy)' column and 'Full Name' column
    timestamps = completed_rows[['Full Name', 'Timestamp (mm/dd/yyyy)']].copy()

    # Convert the timestamps to datetime for sorting
    timestamps['Timestamp (mm/dd/yyyy)'] = pd.to_datetime(timestamps['Timestamp (mm/dd/yyyy)'])

    # Sort by timestamp to get the order of submission
    timestamps = timestamps.sort_values(by='Timestamp (mm/dd/yyyy)').reset_index(drop=True)

    # Normalize the submission order based on first-come, first-serve (earlier submissions get higher scores)
    min_order = 0
    max_order = len(timestamps) - 1

    timestamps['Normalized Score'] = timestamps.index.to_series().apply(
        lambda x: 1 - (x - min_order) / (max_order - min_order))

    # Create the dictionary where key is student name and value is normalized score
    fcfs_scores = dict(zip(timestamps['Full Name'], timestamps['Normalized Score']))

    return fcfs_scores


def get_project_skills():
    # Dictionary of projects and their required skills
    project_skills = {
        "Apples and Oranges Art": ["UI/UX Design", "Machine Learning/AI", "Database"],
        "Cashew": ["Mobile Development (Android, IOS)", "Flutter", "UI/UX Design"],
        "DatasafeHR": ["UI/UX Design", "Data Privacy & Security", "Database"],
        "Daygent": ["Machine Learning/AI", "AI Agentic Systems"],
        "ForOurLastNames": ["Mobile Development (Android, IOS)", "UI/UX Design"],
        "Jascot Development": ["Machine Learning/AI", "Web Development (e.g.,HTML, CSS, JavaScript)", "Database"],
        "Laguna Ocean Foundation": ["UI/UX Design", "Web Development (e.g.,HTML, CSS, JavaScript)"],
        "Makapo": ["UI/UX Design", "Accessibility Design"],
        "MilMentor": ["Web Development (e.g.,HTML, CSS, JavaScript)", "Machine Learning/AI", "UI/UX Design"],
        "ROID INC": ["Mobile Development (Android, IOS)", "UI/UX Design", "Machine Learning/AI"],
        "Skystart App/MVP": ["Database", "Web Development (e.g.,HTML, CSS, JavaScript)"],
        "SVMMARY": ["Machine Learning/AI", "Web Development (e.g.,HTML, CSS, JavaScript)", "UI/UX Design"],
        "SuperNova Academy": ["Database", "Machine Learning/AI", "AR/VR", "UI/UX Design", "Game Development (Unity or Unreal Engine)", "Python", "React"],
        "Balnce.ai": ["Machine Learning/AI", "AI Agentic Systems", "UI/UX Design"],
        "UCI School of Education - Zhenyao Cai": ["Web Development (e.g.,HTML, CSS, JavaScript)", "Database", "Machine Learning/AI"],
        "UCI Social Ecology - Richard Matthew": ["Web Development (e.g.,HTML, CSS, JavaScript)", "Database", "UI/UX Design"]
    }

    return project_skills


def shorten_skill_name(skill_name):
    # Use regular expression to extract the string after the last period
    match = re.search(r'expertise\.\s*(.*)', skill_name)
    return match.group(1) if match else skill_name


def get_skill_rating_columns_with_shortened_names(file_path):
    # Load the Excel file
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")

    # Filter rows that contain "Completed"
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Filter columns that contain "Rate your skills" in the name
    skill_columns = [col for col in completed_rows.columns if "Rate your skills" in col]

    # Extract student names
    student_names = completed_rows['Full Name'].values

    # Create a dictionary where the key is the student name and the value is another dictionary
    # with shortened skill names as keys and corresponding ratings as values
    student_skill_ratings = {}
    for idx, student in enumerate(student_names):
        skills = {
            shorten_skill_name(skill): int(completed_rows.iloc[idx][skill]) if isinstance(completed_rows.iloc[idx][skill], np.float64) else completed_rows.iloc[idx][skill]
            for skill in skill_columns
        }
        student_skill_ratings[student] = skills

    return student_skill_ratings


def get_skill_score(file_path):
    # Step 1: Get the project skills
    project_skills = get_project_skills()

    # Step 2: Get student skill ratings
    student_skill_ratings = get_skill_rating_columns_with_shortened_names(file_path)

    # Create a dictionary where the key is the student name and the value is a dictionary
    # with project names as keys and their respective skill scores as values
    student_skill_scores = {}

    # Iterate over each student
    for student, skills in student_skill_ratings.items():
        project_scores = {}

        # Iterate over each project
        for project, required_skills in project_skills.items():
            # Calculate the sum of ratings for the required skills of the project
            total_score = 0
            count = 0

            for required_skill in required_skills:
                if required_skill in skills:
                    total_score += skills[required_skill]
                    count += 1

            # Normalize the score if there are required skills
            if count > 0:
                normalized_score = total_score / (count * 10)  # Normalizing based on max score of 10
            else:
                normalized_score = 0  # If no matching skills, set score to 0

            project_scores[project] = normalized_score

        # Assign the project scores to the student
        student_skill_scores[student] = project_scores

    return student_skill_scores


def summarize_scores(file_path):
    # Calculate skill scores and project preferences
    skill_scores = get_skill_score(file_path)
    preference_scores = calculate_project_preference(file_path)

    # Create a summary dictionary
    student_summary = {}

    # Combine skill and preference scores for each student
    for student in skill_scores:
        if student in preference_scores:
            combined_scores = {
                project: (skill_scores[student][project]*0.35) + (preference_scores[student][project]*0.65)
                for project in skill_scores[student]}
            student_summary[student] = combined_scores

    return student_summary


def get_grouped_members(file_path):
    # Load the 'Raw Data' sheet for processing
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")

    # Filter rows that contain "Completed"
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Create a dictionary to store the final groupings
    grouped_members = []

    # Create a map from UCI Student ID to Full Name for lookups
    id_to_name = dict(zip(completed_rows["UCI Student ID"], completed_rows["Full Name"]))

    # Process each student
    for index, row in completed_rows.iterrows():
        student_name = row["Full Name"]
        student_id = row["UCI Student ID"]

        # Check if the student wants to team up
        wants_teammates = row.get("Do you have any students you want to team up with?", "").lower() == "yes"

        if not wants_teammates:
            continue

        # Get the number of teammates they want
        try:
            num_teammates = int(row.get("How many students you want to team up with?", 0))
        except (ValueError, TypeError):
            num_teammates = 0

        # Initialize a set for this student's group
        student_group = {student_name}

        # Look up each requested teammate
        for i in range(1, num_teammates + 1):
            member_id_col = f"Member {i} UCI ID"

            if member_id_col in row and pd.notna(row[member_id_col]):
                teammate_id = row[member_id_col]

                # Find the teammate's name using their UCI ID
                if teammate_id in id_to_name:
                    # Get the row for this teammate
                    teammate_rows = completed_rows[completed_rows["UCI Student ID"] == teammate_id]

                    # Only add the teammate if they also want teammates
                    if not teammate_rows.empty:
                        teammate_row = teammate_rows.iloc[0]
                        teammate_wants_teammates = teammate_row.get(
                            "Do you have any students you want to team up with?", "").lower() == "yes"

                        if teammate_wants_teammates:
                            teammate_name = id_to_name[teammate_id]
                            student_group.add(teammate_name)

        # Only add groups with more than one member
        if len(student_group) > 1:
            merge_groups(grouped_members, student_group)

    # Convert grouped_members to a more readable format (lists of students in each group)
    grouped_members_list = [list(group) for group in grouped_members]

    # Save the grouped members to a text file
    with open("grouped_members.txt", "w") as f:
        for group in grouped_members_list:
            f.write(", ".join(group) + "\n")

    return grouped_members_list


def merge_groups(grouped_members, new_members):
    # Find groups that have any overlap with new_members
    overlapping_groups = []
    for group in grouped_members:
        if group.intersection(new_members):
            overlapping_groups.append(group)

    # If no overlapping groups, add the new_members as a new group
    if not overlapping_groups:
        # Check if new_members already exceeds maximum size
        if len(new_members) > 5:
            # Split the group in half
            members_list = list(new_members)
            half_size = len(members_list) // 2
            grouped_members.append(set(members_list[:half_size]))
            grouped_members.append(set(members_list[half_size:]))
        else:
            grouped_members.append(new_members)
        return

    # If there are overlapping groups, merge them with new_members
    merged_group = new_members.copy()
    for group in overlapping_groups:
        merged_group.update(group)
        grouped_members.remove(group)

    # Check if the merged group exceeds the maximum size
    if len(merged_group) > 5:
        # Split the group in half
        members_list = list(merged_group)
        half_size = len(members_list) // 2
        grouped_members.append(set(members_list[:half_size]))
        grouped_members.append(set(members_list[half_size:]))
    else:
        # Add the merged group
        grouped_members.append(merged_group)


def get_least_preferred_projects(file_path):
    # Load the 'Raw Data' sheet for processing
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")

    # Filter rows that contain "Completed"
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Extract the project columns
    project_columns = [col for col in completed_rows.columns if
                       any(key in col for key in project_columns_map_corrected.keys())]

    # Map the project columns to their actual project names
    project_names = [project_columns_map_corrected[col] for col in project_columns_map_corrected.keys() if
                     col in project_columns]

    # Extract project rankings from the completed rows
    project_ranks = completed_rows[project_columns]

    # Sum up the rankings for each project (higher sum = less preferred)
    project_preference_sums = project_ranks.sum(axis=0)

    # Create a dictionary of project names and their total preferences
    project_preference_totals = dict(zip(project_names, project_preference_sums))

    # Sort the projects by their summed preferences in descending order
    sorted_projects_by_preference = sorted(project_preference_totals.items(), key=lambda x: x[1], reverse=True)

    # Return the least preferred projects
    return sorted_projects_by_preference[:1]


def allocate_projects_excluding_least_preferred(file_path):
    # Step 1: Exclude the least preferred projects
    least_preferred_projects = get_least_preferred_projects(file_path)
    excluded_projects = [project for project, _ in least_preferred_projects]

    # Step 2: Allocate projects for grouped members excluding the least preferred ones
    grouped_members = get_grouped_members(file_path)
    print(grouped_members)
    summarized_scores = summarize_scores(file_path)

    # Filter out excluded projects from available projects
    available_projects = {project: 5 for project in summarized_scores[next(iter(summarized_scores))].keys() if
                          project not in excluded_projects}

    # Track projects that can go up to 6
    projects_up_to_6_assigned = 0
    max_projects_up_to_6 = 0

    # Function to calculate the average score of a project for a group
    def calculate_group_project_score(group, project):
        total_score = 0
        count = 0
        for member in group:
            if member in summarized_scores and project in summarized_scores[member]:
                total_score += summarized_scores[member][project]
                count += 1
        return total_score / count if count > 0 else 0

    # Allocate projects to groups
    project_allocations = {}
    for group in grouped_members:
        best_project = None
        best_score = 0

        # Calculate the best project for the group based on summarized scores
        for project, capacity in available_projects.items():
            if capacity >= len(group):  # Ensure the project has enough capacity
                group_score = calculate_group_project_score(group, project)
                if group_score > best_score:
                    best_score = group_score
                    best_project = project

        # Assign the group to the best project
        if best_project:
            project_allocations[tuple(group)] = best_project
            available_projects[best_project] -= len(group)  # Reduce the capacity of the assigned project

            # If the project has reached 6 students, mark it as full
            if available_projects[best_project] == 0 and projects_up_to_6_assigned < max_projects_up_to_6:
                available_projects[best_project] = 1  # Allow 1 more student to make it 6 total
                projects_up_to_6_assigned += 1

    # Step 3: Assign remaining students dynamically
    # List of remaining students who are not in any group
    all_students = set(summarized_scores.keys())
    grouped_students = set(student for group in project_allocations.keys() for student in group)
    remaining_students = all_students - grouped_students

    # Get timestamp info for the remaining students
    # Load the 'Raw Data' sheet for timestamp information
    raw_data = pd.read_excel(file_path, sheet_name="Raw Data")
    completed_rows = raw_data[raw_data.apply(lambda row: row.astype(str).str.contains("Completed").any(), axis=1)]

    # Create a timestamp dictionary for all students
    timestamp_col = "Timestamp"  # Change this to match your actual timestamp column
    student_timestamps = dict(zip(completed_rows["Full Name"], completed_rows[timestamp_col]))

    # Sort remaining students by their timestamp (first-come first-serve)
    remaining_students_with_ts = [(student, student_timestamps.get(student)) for student in remaining_students]
    remaining_students_with_ts.sort(key=lambda x: x[1])  # Sort by timestamp

    # Get just the sorted list of students
    sorted_remaining_students = [student for student, _ in remaining_students_with_ts]

    # Function to find the best project for a student based on top scores and preference
    def find_best_project_for_student(student):
        # Get scores for all available projects with capacity
        project_scores = []
        for project, capacity in available_projects.items():
            if capacity > 0:
                if student in summarized_scores and project in summarized_scores[student]:
                    score = summarized_scores[student][project]
                    project_scores.append((project, score))

        if not project_scores:
            return None

        # Sort by score in descending order and get top 3
        project_scores.sort(key=lambda x: x[1], reverse=True)
        top_projects = project_scores[:3] if len(project_scores) >= 3 else project_scores

        # Extract preference scores from summarized_scores by using skill_scores
        # This requires accessing the original preference calculation
        preference_data = calculate_project_preference(file_path)

        # From the top projects, find the one with highest preference score
        best_project = None
        highest_preference = -1

        for project, _ in top_projects:
            if student in preference_data and project in preference_data[student]:
                preference = preference_data[student][project]
                if preference > highest_preference:
                    highest_preference = preference
                    best_project = project

        # If no preference data is found, fall back to the highest combined score
        if best_project is None and project_scores:
            best_project = project_scores[0][0]

        return best_project

    # Assign remaining students in timestamp order (first-come first-serve)
    for student in sorted_remaining_students:
        best_project = find_best_project_for_student(student)
        if best_project:
            # Assign the student to the project
            found_group = False
            for group in list(project_allocations.keys()):
                if project_allocations[group] == best_project and len(group) < 5:  # Ensure we don't exceed 5 students
                    # Convert tuple group to a list, append student, then convert back to tuple
                    new_group = list(group)
                    new_group.append(student)
                    del project_allocations[group]
                    project_allocations[tuple(new_group)] = best_project
                    found_group = True
                    break
            # If no group was found, create a new one
            if not found_group:
                project_allocations[(student,)] = best_project
            # Reduce the capacity of the assigned project
            available_projects[best_project] -= 1

    return project_allocations


def format_allocations(allocations):
    # Create a dictionary where the project name is the key and members are listed under it
    formatted_output = {}

    for group, project in allocations.items():
        if project not in formatted_output:
            formatted_output[project] = []
        formatted_output[project].extend(group)

    return formatted_output


def preprocess_survey_data(survey_path, students_path, output_dir="."):
    """
    Preprocesses survey data by:
    1. Finding and removing duplicate responses, keeping only the latest one
    2. Identifying students who took the survey but dropped the course
    3. Identifying students in the course who did not take the survey
    4. Updates the original Survey.xlsx with a new sheet containing deduplicated data

    Parameters:
    survey_path (str): Path to the Survey.xlsx file
    students_path (str): Path to the Students.csv file
    output_dir (str): Directory to save output text files

    Returns:
    pd.DataFrame: Preprocessed survey data with duplicates removed
    """
    # Create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    # Load the survey data
    survey_data = pd.read_excel(survey_path, sheet_name="Raw Data")

    # Convert timestamp to datetime for proper sorting
    survey_data["Timestamp"] = pd.to_datetime(survey_data["Timestamp (mm/dd/yyyy)"])

    # Load the student roster
    student_roster = pd.read_csv(students_path)

    # Filter to keep only completed responses
    completed_data = survey_data[survey_data["Response Status"] == "Completed"]

    # Identify incomplete responses
    incomplete_data = survey_data[survey_data["Response Status"] != "Completed"]

    # Step 1: Handle duplicate responses
    # Sort by timestamp in descending order to get the latest responses first
    completed_data = completed_data.sort_values("Timestamp", ascending=False)

    # Find duplicate entries based on email
    duplicates = completed_data[completed_data.duplicated(subset=["UCI Email Address"], keep="first")]

    # Keep only the latest response for each student
    deduplicated_data = completed_data.drop_duplicates(subset=["UCI Email Address"], keep="first")

    # Create a duplicate responses report
    if not duplicates.empty:
        duplicate_report = "Duplicate Survey Responses (Older entries removed):\n\n"
        for _, row in duplicates.iterrows():
            duplicate_report += f"Email: {row['UCI Email Address']}\n"
            duplicate_report += f"Name: {row.get('Full Name', 'N/A')}\n"
            duplicate_report += f"Submission Time: {row['Timestamp']}\n"
            duplicate_report += "-" * 50 + "\n"

        # Save the duplicate report
        with open(os.path.join(output_dir, "duplicate_responses.txt"), "w") as f:
            f.write(duplicate_report)

    # Step 2: Check for students who took the survey but dropped the course
    # And students who are in the course but didn't take the survey

    # Get sets of emails for comparison
    survey_emails = set(deduplicated_data["UCI Email Address"].str.lower())
    roster_emails = set(student_roster["Email"].str.lower())

    # Find students who took the survey but aren't in the roster
    dropped_students = survey_emails - roster_emails

    # Find students in the roster who didn't take the survey
    missing_students = roster_emails - survey_emails

    # Create a report for students who took the survey but dropped and those who didn't take it
    enrollment_report = "Survey Enrollment Report:\n\n"

    if dropped_students:
        enrollment_report += "Students who took the survey but dropped the course:\n"
        for email in dropped_students:
            student_data = deduplicated_data[deduplicated_data["UCI Email Address"].str.lower() == email]
            if not student_data.empty:
                name = student_data.iloc[0].get("Full Name", "Name not provided")
                enrollment_report += f"- {name} ({email})\n"
        enrollment_report += "\n" + "-" * 50 + "\n\n"

    if missing_students:
        enrollment_report += "Students in the course who did not take the survey:\n"
        for email in missing_students:
            student_data = student_roster[student_roster["Email"].str.lower() == email]
            if not student_data.empty:
                name = f"{student_data.iloc[0].get('First Name', '')} {student_data.iloc[0].get('Last Name', '')}"
                enrollment_report += f"- {name.strip()} ({email})\n"

    # If there were incomplete responses, add them to the report
    if not incomplete_data.empty:
        enrollment_report += "\n" + "-" * 50 + "\n\n"
        enrollment_report += "Students with incomplete survey responses (excluded from analysis):\n"
        for _, row in incomplete_data.iterrows():
            email = row.get("UCI Email Address", "No email provided")
            name = row.get("Full Name", "Name not provided")
            enrollment_report += f"- {name} ({email})\n"

    # Save the enrollment report
    with open(os.path.join(output_dir, "enrollment_report.txt"), "w") as f:
        f.write(enrollment_report)

    # Step 3: Update the original Survey.xlsx file by replacing "Raw Data" sheet
    try:
        # First, save a backup of the original file
        backup_path = survey_path.replace(".xlsx", "_backup.xlsx")
        shutil.copy2(survey_path, backup_path)
        print(f"Created backup of original file at {backup_path}")

        # Get all sheets from the workbook
        book = load_workbook(survey_path)
        sheet_names = book.sheetnames

        # Create a new Excel writer
        with pd.ExcelWriter(survey_path, engine='openpyxl') as writer:
            # Replace the "Raw Data" sheet with deduplicated data
            deduplicated_data.to_excel(writer, sheet_name="Raw Data", index=False)

            # Copy all other sheets from the original workbook
            original_wb = pd.ExcelFile(backup_path)
            for sheet_name in sheet_names:
                if sheet_name != "Raw Data":
                    df = pd.read_excel(original_wb, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Successfully updated {survey_path} by replacing 'Raw Data' with deduplicated data")
    except Exception as e:
        print(f"Error updating the Excel file: {e}")
        print("Saving deduplicated data to a separate file instead")
        deduplicated_data.to_excel(os.path.join(output_dir, "Deduplicated_Survey.xlsx"), index=False)

    print(f"Preprocessing complete:")
    print(f"- Found {len(duplicates)} duplicate responses")
    print(f"- Found {len(incomplete_data)} incomplete responses")
    print(f"- Found {len(dropped_students)} students who took the survey but dropped the course")
    print(f"- Found {len(missing_students)} students in the course who didn't take the survey")
    print(f"Reports saved to {output_dir}/duplicate_responses.txt and {output_dir}/enrollment_report.txt")


def create_allocation_excel(formatted_allocations, survey_path, project_columns_map,
                            output_path="Project_Allocations.xlsx"):
    """
    Creates an Excel file with project allocations, including each student's ranking for their assigned project.

    Parameters:
    formatted_allocations (dict): Dictionary mapping project names to lists of student names
    survey_path (str): Path to the survey Excel file
    project_columns_map (dict): Mapping from column names to project names
    output_path (str): Path for the output Excel file
    """
    # Load the survey data
    survey_data = pd.read_excel(survey_path, sheet_name="Raw Data")

    # Create reverse mapping from project names to column names
    project_to_column = {v: k for k, v in project_columns_map.items()}

    # Get student rankings for each project
    student_project_rankings = {}

    # Process each project and its assigned students
    for project, students in formatted_allocations.items():
        if project not in project_to_column:
            print(f"Warning: Project '{project}' not found in project_columns_map")
            continue

        column_name = project_to_column[project]
        student_rankings = []

        for student in students:
            # Find the student in the survey data
            student_row = survey_data[survey_data["Full Name"] == student]

            if student_row.empty:
                # Try to find by partial match
                for index, row in survey_data.iterrows():
                    full_name = str(row.get("Full Name", ""))
                    if student.lower() in full_name.lower():
                        student_row = survey_data.iloc[[index]]
                        break

            if student_row.empty:
                print(f"Warning: Could not find data for student '{student}'")
                ranking = "N/A"
            else:
                # Get the student's ranking for this project
                try:
                    ranking = student_row[column_name].values[0]
                    # Check if ranking is a number
                    if pd.isna(ranking) or not np.issubdtype(type(ranking), np.number):
                        ranking = "N/A"
                except:
                    ranking = "N/A"

            # Add student with ranking
            student_rankings.append(f"{student} ({ranking})")

        student_project_rankings[project] = student_rankings

    # Find the maximum number of students in any project
    max_students = max(len(students) for students in student_project_rankings.values())

    # Create a DataFrame with projects as columns and students as rows
    data = {}
    for project, students_with_rankings in student_project_rankings.items():
        # Pad with empty strings if necessary to ensure all columns have the same number of rows
        padded_students = students_with_rankings + [''] * (max_students - len(students_with_rankings))
        data[project] = padded_students

    # Create DataFrame
    df = pd.DataFrame(data)

    # Sort columns alphabetically by project name
    df = df.reindex(sorted(df.columns), axis=1)

    # Write to Excel
    df.to_excel(output_path, index=False)

    print(f"Project allocations with rankings saved to {output_path}")


# Preprocess the survey data - this will update the original Excel file
preprocess_survey_data("Survey.xlsx", "Students.csv", "output_reports")
# Execute the combined function to allocate projects to both grouped and remaining students
final_project_allocations = allocate_projects_excluding_least_preferred("Survey.xlsx")

# Format the final allocations for display
formatted_final_allocations = format_allocations(final_project_allocations)
create_allocation_excel(
    formatted_allocations=formatted_final_allocations,
    survey_path="Survey.xlsx",
    project_columns_map=project_columns_map_corrected,
    output_path="Project_Allocations.xlsx"
)